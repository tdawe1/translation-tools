
import openpyxl
from openpyxl.utils.cell import range_boundaries

xlsx_path = 'inputs/693650c6b8725_7800b1bed3ec11f0910c8a06568e45ec.xlsx'
wb = openpyxl.load_workbook(xlsx_path)
sheet = wb.active

comments_to_add = {
    19: [
        "いただいたご指摘のうち、2点目と4点目につきましては、一度操作を行えば文脈から自然に理解される内容でしたので、過剰な説明を避ける目的で短縮しておりました。文脈理解を前提とした設計上、簡潔な表現のほうが自然であると判断したためです。",
        "また、日本語は英語に比べて1文字あたりに含まれる情報量が多く、表示スペースの制約も異なるため、英語版ではより要約的な表現を採用する傾向があります。今回の調整はそうした一般的なUI設計上の判断に基づくものです。"
    ],
    42: [
        "英語において「self-consciousness」は自我（自意識）を表す最も本質的な語であり、「self-awareness」とは使われる文脈や思想的背景が異なります。モデルが「self-conscious（自意識過剰な）」という形容詞と混同している可能性も考えられます。",
        "哲学的・SF的な文脈では「self-consciousness」がより正確で深みのある表現と考えております。もし変更をご希望であれば対応可能ですが、現在の訳が最も適切であると判断いたします。"
    ]
}

# Map merged cells to their master cell
merged_cells_map = {}
for merged_range in sheet.merged_cells.ranges:
    min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
    master_cell_coords = (min_row, min_col)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            merged_cells_map[(r, c)] = master_cell_coords

rows_updated_count = 0
for row_num, comments_list in comments_to_add.items():
    col_num = 5 # Column E
    
    target_coords = (row_num, col_num)
    if target_coords in merged_cells_map:
        master_row, master_col = merged_cells_map[target_coords]
        comment_cell = sheet.cell(row=master_row, column=master_col)
    else:
        comment_cell = sheet.cell(row=row_num, column=col_num)

    current_comment = comment_cell.value if comment_cell.value else ""
    new_comments_str = "\n\n".join(comments_list)
    
    if current_comment:
        comment_cell.value = current_comment + "\n\n" + new_comments_str
    else:
        comment_cell.value = new_comments_str
    rows_updated_count += 1
    print(f"Updated Row {row_num}, Column E with comments.")

wb.save(xlsx_path)
print(f"Finished. Total rows updated: {rows_updated_count}.")
print(f"Saved to {xlsx_path}")
