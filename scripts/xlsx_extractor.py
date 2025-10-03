from openpyxl import load_workbook
from typing import List, Tuple, Dict, Any, Optional
from .base_extractor import BaseExtractor

class XlsxExtractor(BaseExtractor):
    def extract(self, file_path: str, page_range: Optional[str] = None) -> List[Tuple[str, int, str, Dict[str, Any]]]:
        wb = load_workbook(file_path, data_only=True)
        blocks = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                row_idx = ws.max_row if not row else list(row).index(row) + 1  # Approximate
                for col_idx, cell_value in enumerate(row, 1):
                    if isinstance(cell_value, str) and cell_value.strip():
                        text = cell_value.strip()
                        section = sheet_name
                        index = f'{row_idx},{col_idx}'  # Use coordinate as index
                        layout = {
                            'type': 'cell',
                            'sheet': sheet_name,
                            'coordinate': f'{chr(64 + col_idx)}{row_idx}',
                            'font_size': 11.0  # Default, can read from cell.font if needed
                        }
                        blocks.append((section, index, text, layout))
        
        return blocks